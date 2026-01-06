from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(80, max(10, max_len + 2))


def _ordered_headers(rows: list[dict[str, object]]) -> list[str]:
    if not rows:
        return []

    keys = set()
    for r in rows:
        keys.update(r.keys())

    preferred = ["index", "model", "queries", "winner", "error"]
    cols: list[str] = [k for k in preferred if k in keys]

    rest = sorted([k for k in keys if k not in cols])
    cols.extend(rest)
    return cols


def _write_sheet(ws, rows: list[dict[str, object]]) -> None:
    headers = _ordered_headers(rows)
    if not headers:
        ws.append(["_empty_"])
        return
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    _autosize(ws)


def write_summary_xlsx(
    path: str,
    summary_rows: list[dict[str, object]],
    per_index_sheets: dict[str, list[dict[str, object]]],
    extra_sheets: dict[str, list[dict[str, object]]] | None = None,
) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _write_sheet(ws, summary_rows)

    # Optional: extra top-level sheets like Delta
    if extra_sheets:
        for sheet_name, rows in extra_sheets.items():
            safe = sheet_name[:31]
            wsx = wb.create_sheet(title=safe)
            _write_sheet(wsx, rows)

    # Per-index detail sheets
    for sheet_name, rows in per_index_sheets.items():
        safe = sheet_name[:31]
        ws2 = wb.create_sheet(title=safe)
        _write_sheet(ws2, rows)

    wb.save(path)


def render_summary_table_md(rows: list[dict[str, object]]) -> str:
    if not rows:
        return "_No results_"

    cols = _ordered_headers(rows)
    if not cols:
        return "_No results_"

    header = "| " + " | ".join(cols) + " |"
    sep = "| " + " | ".join(["---"] * len(cols)) + " |"

    lines = [header, sep]
    for r in rows:
        lines.append("| " + " | ".join(str(r.get(c, "")) for c in cols) + " |")
    return "\n".join(lines)
